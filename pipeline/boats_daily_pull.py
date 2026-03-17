"""
boats_daily_pull.py - Automated BOATS (BlueOcean ATS) daily data pull
=====================================================================
Downloads Market_Data_Statistics xlsx files from email.

Pull methods (in priority order):
  1. Gmail IMAP  - Search forwarded BOATS emails in Gmail (primary)
  2. Graph API   - Search CalcGuard Outlook mailbox via Microsoft Graph (fallback)

Emails arrive from Matthew Brown at BlueOcean ATS with subject pattern
"boMMDD" (e.g., "bo0313" for Mar 13). The CalcGuard Outlook rule forwards
these to Gmail automatically.

Usage:
    python boats_daily_pull.py                     # Pull today's data
    python boats_daily_pull.py --date 2026-03-13   # Pull specific date
    python boats_daily_pull.py --check-gaps         # Report missing files
    python boats_daily_pull.py --method gmail       # Force Gmail only
    python boats_daily_pull.py --method graph       # Force Graph API only

Exit codes:
    0 = Success (file saved)
    1 = Skipped (weekend, holiday, file exists, or email not yet received)
    2 = Error (auth failure, API error, validation failure)

Prerequisites:
    pip install msal requests openpyxl
    Configure .env with:
      Gmail:  GMAIL_ADDRESS, GMAIL_APP_PASSWORD
      Graph:  GRAPH_TENANT_ID, GRAPH_CLIENT_ID, GRAPH_CLIENT_SECRET, GRAPH_USER_EMAIL

Sapinover LLC | Overnight Trading Research Platform
"""

import os
import sys
import io
import re
import time
import base64
import email
import imaplib
import argparse
import datetime

import requests

try:
    import msal
    _HAS_MSAL = True
except ImportError:
    _HAS_MSAL = False

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(2)

from pull_utils import (
    BASE_DIR, is_trading_day, setup_logging,
    dual_save_bytes, load_env
)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

GRAPH_BASE_URL = "https://graph.microsoft.com/v1.0"
SENDER_EMAIL = "matthew.brown@blueoceanats.com"
MAX_RETRIES = 3

# Expected BOATS xlsx columns (11 required; VWAP optional in some files)
EXPECTED_COLUMNS = {
    "Symbol", "Notional", "Volume", "Executions",
    "Open", "Hi", "Lo", "Close",
    "Average Spread", "Average Bid Size", "Average Offer Size"
}

logger = None  # Initialized in main()


# ===========================================================================
# METHOD 1: Gmail IMAP (PRIMARY)
# ===========================================================================

def gmail_pull(env: dict, target_date: datetime.date) -> tuple | None:
    """Search Gmail via IMAP for the forwarded BOATS email, download xlsx.
    Returns (file_bytes, original_filename) or None."""

    gmail_addr = env.get("GMAIL_ADDRESS", "")
    gmail_pass = env.get("GMAIL_APP_PASSWORD", "")

    if not gmail_addr or not gmail_pass:
        logger.debug("Gmail credentials not configured (GMAIL_ADDRESS, GMAIL_APP_PASSWORD)")
        return None

    mmdd = target_date.strftime("%m%d")
    subject_pattern = f"bo{mmdd}"

    logger.info(f"[Gmail IMAP] Searching for subject '{subject_pattern}' in {gmail_addr}...")

    conn = None
    try:
        # Connect to Gmail IMAP
        conn = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        conn.login(gmail_addr, gmail_pass)
        conn.select("INBOX")

        # Search for emails with the BOATS subject pattern
        # IMAP SUBJECT search is substring match
        status, msg_ids = conn.search(None, f'(SUBJECT "{subject_pattern}")')

        if status != "OK" or not msg_ids[0]:
            # Also try searching with "Fwd:" prefix in case forwarded
            status, msg_ids = conn.search(None, f'(SUBJECT "Fwd: {subject_pattern}")')

        if status != "OK" or not msg_ids[0]:
            # Try broader search: subject contains the date pattern
            # and from contains blueocean or the forwarding address
            status, msg_ids = conn.search(None, f'(SUBJECT "{subject_pattern}")')

        if status != "OK" or not msg_ids[0]:
            logger.info(f"  No email found with subject containing '{subject_pattern}'")
            return None

        # Get the most recent matching message
        id_list = msg_ids[0].split()
        latest_id = id_list[-1]  # Most recent

        # Fetch the full email
        status, msg_data = conn.fetch(latest_id, "(RFC822)")
        if status != "OK":
            logger.error("  Failed to fetch email content")
            return None

        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)

        logger.info(f"  Found: \"{msg['Subject']}\" from {msg['From']} ({msg['Date']})")

        # Walk through parts looking for xlsx attachment
        for part in msg.walk():
            content_disp = str(part.get("Content-Disposition", ""))
            if "attachment" not in content_disp:
                continue

            filename = part.get_filename()
            if not filename:
                continue

            # Decode filename if encoded
            if isinstance(filename, bytes):
                filename = filename.decode("utf-8", errors="replace")

            # Clean up any RFC2047 encoding
            decoded_parts = email.header.decode_header(filename)
            filename = ""
            for decoded_str, charset in decoded_parts:
                if isinstance(decoded_str, bytes):
                    filename += decoded_str.decode(charset or "utf-8", errors="replace")
                else:
                    filename += decoded_str

            if filename.lower().endswith(".xlsx") and "Market_Data_Statistics" in filename:
                file_bytes = part.get_payload(decode=True)
                if file_bytes:
                    logger.info(f"  Downloaded attachment: {filename} ({len(file_bytes):,} bytes)")
                    return (file_bytes, filename)
                else:
                    logger.error(f"  Attachment {filename} has empty payload")
                    return None

        # No xlsx found, check what attachments exist
        att_names = []
        for part in msg.walk():
            fn = part.get_filename()
            if fn:
                att_names.append(fn)

        if att_names:
            logger.warning(f"  No Market_Data_Statistics xlsx found. Attachments: {att_names}")
        else:
            logger.warning("  Email found but has no attachments")

        return None

    except imaplib.IMAP4.error as e:
        error_str = str(e)
        if "AUTHENTICATIONFAILED" in error_str or "Invalid credentials" in error_str:
            logger.error(f"  Gmail auth failed. Check GMAIL_APP_PASSWORD in .env")
            logger.error(f"  Generate app password at: https://myaccount.google.com/apppasswords")
        else:
            logger.error(f"  Gmail IMAP error: {e}")
        return None

    except Exception as e:
        logger.error(f"  Gmail error: {e}")
        return None

    finally:
        if conn:
            try:
                conn.close()
                conn.logout()
            except Exception:
                pass


# ===========================================================================
# METHOD 2: Microsoft Graph API (FALLBACK)
# ===========================================================================

def get_graph_token(env: dict) -> str | None:
    """Acquire OAuth2 token via MSAL client_credentials flow."""
    if not _HAS_MSAL:
        logger.debug("msal not installed, skipping Graph API")
        return None

    tenant_id = env.get("GRAPH_TENANT_ID", "")
    client_id = env.get("GRAPH_CLIENT_ID", "")
    client_secret = env.get("GRAPH_CLIENT_SECRET", "")

    if not all([tenant_id, client_id, client_secret]):
        logger.debug("Missing Graph API credentials in .env")
        return None

    authority = f"https://login.microsoftonline.com/{tenant_id}"
    scopes = ["https://graph.microsoft.com/.default"]

    app = msal.ConfidentialClientApplication(
        client_id=client_id,
        client_credential=client_secret,
        authority=authority
    )

    result = app.acquire_token_for_client(scopes=scopes)

    if "access_token" in result:
        logger.debug("Graph API token acquired (client_credentials flow)")
        return result["access_token"]

    error = result.get("error_description", result.get("error", "Unknown error"))
    logger.warning(f"  Graph token failed: {error}")
    return None


def _graph_get(url: str, token: str, params: dict = None) -> dict | None:
    """Make a GET request to Graph API with retry logic."""
    headers = {"Authorization": f"Bearer {token}"}

    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            if resp.status_code == 401:
                logger.error("Graph API returned 401 Unauthorized.")
                return None
            if resp.status_code == 403:
                logger.warning("Graph API returned 403 Forbidden (admin consent not granted).")
                return None
            resp.raise_for_status()
            return resp.json()
        except requests.RequestException as e:
            wait = 2 ** (attempt + 1)
            logger.warning(f"  Attempt {attempt + 1}/{MAX_RETRIES} failed: {e}. Retrying in {wait}s...")
            time.sleep(wait)

    logger.error(f"Graph API request failed after {MAX_RETRIES} attempts")
    return None


def graph_pull(env: dict, target_date: datetime.date) -> tuple | None:
    """Search Outlook mailbox via Graph API for BOATS email.
    Returns (file_bytes, original_filename) or None."""

    user_email = env.get("GRAPH_USER_EMAIL", "")
    if not user_email:
        logger.debug("GRAPH_USER_EMAIL not set, skipping Graph API")
        return None

    token = get_graph_token(env)
    if token is None:
        return None

    mmdd = target_date.strftime("%m%d")
    subject_prefix = f"bo{mmdd}"

    # Search inbox
    filter_str = (
        f"from/emailAddress/address eq '{SENDER_EMAIL}' "
        f"and startswith(subject, '{subject_prefix}')"
    )

    url = f"{GRAPH_BASE_URL}/users/{user_email}/mailFolders/inbox/messages"
    params = {
        "$filter": filter_str,
        "$select": "id,subject,receivedDateTime,hasAttachments",
        "$top": 5,
        "$orderby": "receivedDateTime desc"
    }

    logger.info(f"[Graph API] Searching for email with subject '{subject_prefix}'...")
    data = _graph_get(url, token, params)

    if data is None:
        return None

    messages = data.get("value", [])
    if not messages:
        logger.info(f"  No email found with subject starting with '{subject_prefix}'")
        return None

    msg = messages[0]
    logger.info(f"  Found: \"{msg['subject']}\" received {msg['receivedDateTime']}")

    if not msg.get("hasAttachments", False):
        logger.warning("  Email found but has no attachments.")
        return None

    # Download attachments
    att_url = f"{GRAPH_BASE_URL}/users/{user_email}/messages/{msg['id']}/attachments"
    att_data = _graph_get(att_url, token)

    if att_data is None:
        return None

    attachments = att_data.get("value", [])
    for att in attachments:
        name = att.get("name", "")
        if name.lower().endswith(".xlsx") and "Market_Data_Statistics" in name:
            logger.info(f"  Downloading attachment: {name} ({att.get('size', 0):,} bytes)")
            content_b64 = att.get("contentBytes", "")
            if not content_b64:
                logger.error(f"  Attachment {name} has no contentBytes.")
                return None
            file_bytes = base64.b64decode(content_b64)
            return (file_bytes, name)

    found_names = [a.get("name", "unknown") for a in attachments]
    logger.warning(f"  No .xlsx attachment found. Attachments present: {found_names}")
    return None


# ===========================================================================
# Validation
# ===========================================================================

def validate_boats_xlsx(file_bytes: bytes, target_date: datetime.date) -> bool:
    """Validate the downloaded BOATS xlsx file."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active

        # Check column headers (row 1)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        headers_clean = {h.strip() if isinstance(h, str) else h for h in headers if h}

        if not EXPECTED_COLUMNS.issubset(headers_clean):
            missing = EXPECTED_COLUMNS - headers_clean
            logger.error(f"  Validation failed: missing columns {missing}")
            logger.error(f"  Found columns: {headers}")
            wb.close()
            return False

        # VWAP is optional (some BOATS files have 11 cols without VWAP, some have 12 with VWAP)
        if "VWAP" in headers_clean:
            logger.info(f"  VWAP column present (12-column format)")
        else:
            logger.info(f"  VWAP column absent (11-column format, pipeline will handle)")

        row_count = sum(1 for _ in ws.iter_rows(min_row=2))
        wb.close()

        if row_count < 500:
            logger.warning(f"  Low row count: {row_count} (expected ~9,000). Saving anyway.")
        else:
            logger.info(f"  Validation passed: {row_count:,} rows, {len(headers_clean)} columns")

        size_kb = len(file_bytes) / 1024
        if size_kb < 50:
            logger.warning(f"  Unusually small file: {size_kb:.0f} KB")
        elif size_kb > 5000:
            logger.warning(f"  Unusually large file: {size_kb:.0f} KB")

        return True

    except Exception as e:
        logger.error(f"  Validation failed: {e}")
        return False


# ===========================================================================
# File management
# ===========================================================================

def make_filename(d: datetime.date) -> str:
    """Generate canonical BOATS filename."""
    return f"Market_Data_Statistics_{d.strftime('%Y%m%d')}.xlsx"


def file_already_exists(d: datetime.date) -> bool:
    """Check if BOATS file exists in root, month dir, or BlueOcean/Market_Data/."""
    filename = make_filename(d)
    locations = [
        os.path.join(BASE_DIR, filename),
        os.path.join(BASE_DIR, d.strftime("%Y-%m"), filename),
        os.path.join(BASE_DIR, "BlueOcean", "Market_Data", filename),
    ]
    return any(os.path.exists(p) for p in locations)


def check_for_gaps(lookback_days: int = 5):
    """Report missing BOATS files for recent trading days."""
    today = datetime.date.today()
    missing = []
    for i in range(1, lookback_days + 1):
        check_date = today - datetime.timedelta(days=i)
        if is_trading_day(check_date) and not file_already_exists(check_date):
            missing.append(check_date)

    if missing:
        logger.warning(f"Missing BOATS data for {len(missing)} trading day(s):")
        for d in sorted(missing):
            logger.warning(f"  {d.strftime('%Y-%m-%d')} ({d.strftime('%A')})")
    else:
        logger.info(f"No gaps found in the last {lookback_days} trading days.")


# ===========================================================================
# Main
# ===========================================================================

def main():
    global logger
    logger = setup_logging("boats_pull", "boats_pull.log")

    parser = argparse.ArgumentParser(description="BOATS (BlueOcean ATS) daily data pull")
    parser.add_argument("--date", type=str, help="Target date (YYYY-MM-DD). Default: today")
    parser.add_argument("--check-gaps", action="store_true", help="Check for missing recent files")
    parser.add_argument("--method", choices=["gmail", "graph", "auto"], default="auto",
                        help="Pull method: gmail, graph, or auto (try both)")
    args = parser.parse_args()

    logger.info("=" * 50)
    logger.info("BOATS Daily Pull (Gmail IMAP + Graph API)")
    logger.info("=" * 50)

    # Gap check mode
    if args.check_gaps:
        check_for_gaps()
        return 0

    # Determine target date
    if args.date:
        try:
            target_date = datetime.datetime.strptime(args.date, "%Y-%m-%d").date()
        except ValueError:
            logger.error(f"Invalid date format: {args.date}. Use YYYY-MM-DD.")
            return 2
    else:
        target_date = datetime.date.today()

    logger.info(f"Target date: {target_date.strftime('%Y-%m-%d')} ({target_date.strftime('%A')})")

    # Check if trading day
    if not is_trading_day(target_date):
        reason = "weekend" if target_date.weekday() >= 5 else "market holiday"
        logger.info(f"Skipping: {target_date} is a {reason}.")
        return 1

    # Check if already pulled
    if file_already_exists(target_date):
        logger.info(f"Skipping: {make_filename(target_date)} already exists.")
        return 1

    # Load credentials
    env = load_env()
    result = None

    # --- Try pull methods in priority order ---

    if args.method in ("gmail", "auto"):
        result = gmail_pull(env, target_date)
        if result:
            logger.info("  Source: Gmail IMAP")

    if result is None and args.method in ("graph", "auto"):
        result = graph_pull(env, target_date)
        if result:
            logger.info("  Source: Graph API")

    if result is None:
        if args.method == "auto":
            logger.info("Email not found via Gmail or Graph API. Will retry on next scheduled run.")
        else:
            logger.info(f"Email not found via {args.method}. Will retry on next scheduled run.")
        return 1

    file_bytes, original_name = result

    # Validate
    if not validate_boats_xlsx(file_bytes, target_date):
        logger.error("File validation failed. Not saving.")
        return 2

    # Save to both locations
    filename = make_filename(target_date)
    saved = dual_save_bytes(file_bytes, filename, target_date, logger)

    logger.info(f"Summary: {len(file_bytes):,} bytes, saved to {len(saved)} location(s)")
    logger.info("Completed successfully.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
